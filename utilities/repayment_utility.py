from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment, Color
from openpyxl.styles.proxy import StyleProxy
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.worksheet import Worksheet
from configs.constant_data import ConstData
from models.excel_cell_data import ExcelCellData
from models.loan import Loan
from models.loanee import Loanee
from models.repayment import Repayment
from utilities.excel_utility import ExcelUtility
from utilities.misc_utility import MiscUtility
import re


class RepaymentUtility:
    centered_alignment = Alignment(horizontal='center', vertical='center')

    # @staticmethod
    # def mark_refunded_repayments(repayments: list[Repayment], refunded: dict[str, list[(str, str)]]):
    #
    #     for loanee_id, refunded_list in refunded.items():
    #         corresponding_repayment = [rep for rep in repayments if rep.loanee == loanee_id]
    #
    #         for slice in repayments

    @staticmethod
    def find_paid_slices(file_path: str):
        """
        After repayments have been added to sheets, we look for those which are already refunded : they have a fill color
        :param file_path:
        :return:
        """
        workbook: Workbook | None = None
        paid_slices: dict[str, list[(str, str, int)]] = {}

        try:
            workbook = MiscUtility.open_workbook(file_path)

            for sheet_name in [sh for sh in workbook.sheetnames if re.search(fr"^{ConstData.excel_sheet_repayment}\s*\d+$", sh) is not None]:
                sheet = workbook.get_sheet_by_name(sheet_name)
                year = int(sheet_name.replace(ConstData.excel_sheet_repayment, "").strip())

                for i in range(2, sheet.max_row + 1):
                    row_index = i - 1
                    loanee_id_cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_loan_debtor_ID, row_index)
                    loanee_id = sheet[loanee_id_cell_index].value

                    for month_name in ConstData.months:
                        month_index = ExcelUtility.get_cell_from_column_name(month_name, row_index)
                        slice_cell: Cell = sheet[month_index]

                        if slice_cell.value is None:
                            continue

                        cell_fill: StyleProxy = slice_cell.fill
                        cell_bg: Color = cell_fill.bgColor

                        if cell_bg.value == "00000000":
                            continue

                        if paid_slices.get(loanee_id) is None:
                            paid_slices[loanee_id] = []

                        paid_slices[loanee_id].append((month_name, year, int(slice_cell.value)))

            return paid_slices

        except Exception as e:
            raise e
        finally:
            if workbook is not None:
                MiscUtility.close_workbook(workbook)

    @staticmethod
    def write_repayments_to_excel(file_path: str, data: dict[int, dict[str, ExcelCellData]]):
        """
        Write to file
        :param file_path:
        :param data:
        :return:
        """
        workbook = MiscUtility.open_workbook(file_path)
        sheet_prefix = ConstData.excel_sheet_repayment

        try:
            for year, cells_data in data.items():
                sheet_name = f"{sheet_prefix} {year}"
                sheet = ExcelUtility.write_data_to_sheet(workbook, sheet_name, cells_data, ConstData.excel_cols_repayments)

                RepaymentUtility.stylize_repayment_sheet_specific_columns(sheet)

            workbook.save(file_path)
        except (Exception,):
            raise
        finally:
            MiscUtility.close_workbook(workbook)

    @staticmethod
    def stylize_sheet_general_cells(sheet: Worksheet, columns: tuple[str, ...], start_col: int = None, end_col: int = None, col_width = 20):
        width: int

        start_col = start_col or 0
        end_col = end_col or len(columns)

        # FONTS
        headers_row: RowDimension = sheet.row_dimensions[1]

        # SIZES
        # - Widths
        for index in range(start_col, end_col):
            column_index_letter = ConstData.alphabet[index]

            column_dimension_holder: ColumnDimension = sheet.column_dimensions[column_index_letter]
            column_dimension_holder.width = col_width

            # Headers
            header_cell: Cell = sheet[f"{column_index_letter}1"]

            header_cell.alignment = RepaymentUtility.centered_alignment
            header_cell.font = Font(
                size=12,
                bold=True
            )

        # -- Heights
        headers_row.height = 40

        for i in range(2, sheet.max_row + 1):
            sheet.row_dimensions[i].height = 30

    @staticmethod
    def stylize_repayment_sheet_specific_columns(sheet: Worksheet):
        width: int

        # FONTS
        headers_row: RowDimension = sheet.row_dimensions[1]

        # SIZES
        # - Widths
        start_index = ConstData.excel_cols_repayments.index(ConstData.excel_col_repayment_yearly_amount_loaned)

        for index in range(start_index, len(ConstData.excel_cols_repayments)):
            if index == 3:  # Amount
                width = 30
            else:  # Months
                width = 20

            column_index_letter = ConstData.alphabet[index]

            column_dimension_holder: ColumnDimension = sheet.column_dimensions[column_index_letter]
            column_dimension_holder.width = width

            # # Headers
            # header_cell: Cell = sheet[f"{column_index_letter}1"]
            #
            # header_cell.alignment = RepaymentUtility.centered_alignment
            # header_cell.font = Font(
            #     size=12,
            #     bold=True
            # )

    @staticmethod
    def group_repayments_by_year(repayments: list[Repayment]):
        """
        Simply split and group repayments per year into a dictionary having as key te year and as value repayments
        :param repayments:
        :return:
        """
        repayments_grouped: dict[int, list[Repayment]] = {}

        for index, repayment in enumerate(repayments):
            years = set([sl.date.year for sl in repayment.slices])

            for year in years:
                if repayments_grouped.get(year) is None:
                    repayments_grouped[year] = []

                new_repayment = Repayment(repayment.loanee)
                new_repayment.slices = [sl for sl in repayment.slices if sl.date.year == year]

                repayments_grouped[year].append(new_repayment)

        repayments_grouped = dict(sorted(repayments_grouped.items()))

        return repayments_grouped

    @staticmethod
    def compute_sheets_data(repayments_grouped: dict[int, list[Repayment]]):
        """
        After computing loans to repayment object, this method compute repayments to sheets data into a dictionary
        :param repayments_grouped:
        :return:
        """

        sheets_data: dict[int, dict[str, ExcelCellData]] = {}

        for year, repayments in repayments_grouped.items():
            index = 0

            for repayment in repayments:
                index += 1

                for the_slice in repayment.slices:

                    if sheets_data.get(year) is None:
                        sheets_data[year] = {}
                        RepaymentUtility.fill_headers_cells(sheets_data[year], ConstData.excel_cols_repayments)

                    RepaymentUtility.fill_loanee_cells(sheets_data[year], repayment.loanee, index, ConstData.excel_cols_repayments)
                    RepaymentUtility.fill_annual_total_cells(sheets_data[year], repayment, index, ConstData.excel_cols_repayments)

                    month = ConstData.months[the_slice.date.month - 1]

                    cell_index = ExcelUtility.get_cell_from_column_name(month, index)
                    sheets_data[year][cell_index] = ExcelCellData(the_slice.amount)

        sheets_data = dict(sorted(sheets_data.items()))

        return sheets_data

    # @staticmethod
    # def compute_sheets_data(repayments: list[Repayment]):
    #     """
    #     After computing loans to repayment object, this method compute repayments to sheets data into a dictionary
    #     :param repayments:
    #     :return:
    #     """
    #
    #     sheets_data: dict[int, dict[str, Any]] = {}
    #
    #     for index, repayment in enumerate(repayments):
    #         year = -1
    #
    #         for the_slice in repayment.slices:
    #             if year != the_slice.date.year:
    #                 year = the_slice.date.year
    #
    #                 if sheets_data.get(year) is None:
    #                     sheets_data[year] = {}
    #                     RepaymentUtility.fill_headers_cells(sheets_data, year)
    #
    #                 RepaymentUtility.fill_loanee_cells(sheets_data, year, repayment, index + 1)
    #
    #             month = ConstData.months[the_slice.date.month - 1]
    #
    #             cell_index = ExcelUtility.get_repayment_cell_from_column_name(month, index + 1)
    #             sheets_data[year][cell_index] = the_slice.amount
    #
    #     sheets_data = dict(sorted(sheets_data.items()))
    #
    #     return sheets_data

    @staticmethod
    def fill_headers_cells(sheets_data: dict[str, ExcelCellData], columns: tuple[str, ...]):
        for index, column_name in enumerate(columns):
            sheets_data[f"{ConstData.alphabet[index]}1"] = ExcelCellData(column_name)

    @staticmethod
    def fill_loanee_cells(sheets_data: dict[str, ExcelCellData], loanee: Loanee, index: int, columns: tuple[str, ...]):

        # ID
        cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_loan_debtor_ID, index, columns)
        sheets_data[cell_index] = ExcelCellData(loanee.ID)

        # Firstname
        cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_loan_debtor_first_name, index, columns)
        sheets_data[cell_index] = ExcelCellData(loanee.first_name)

        # Lastname
        cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_loan_debtor_last_name, index, columns)
        sheets_data[cell_index] = ExcelCellData(loanee.last_name)

    @staticmethod
    def fill_annual_total_cells(sheets_data: dict[str, ExcelCellData], repayment: Repayment, index: int, columns: tuple[str, ...]):
        # Total
        cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_repayment_yearly_amount_loaned, index, columns)
        sheets_data[cell_index] = ExcelCellData(sum([slice.amount for slice in repayment.slices]), Font(color="104db0", bold=False))

    @staticmethod
    def compute_repayments(loans: list[Loan]) -> list[Repayment]:
        """
        From provided loans, generate corresponding repayments with slices and dates
        :param loans:
        :return:
        """

        repayments_repository: dict[str, Repayment] = {}

        for loan in loans:
            repayment = repayments_repository.get(loan.loanee.ID)

            if not repayment:
                repayment = Repayment(loan.loanee)
                repayments_repository[loan.loanee.ID] = repayment

            repayment.add_slices(loan.repayment_logic, loan.date)

        return list(repayments_repository.values())

    @staticmethod
    def write_stats_to_excel(data: dict[str, ExcelCellData], file_path: str):
        workbook = MiscUtility.open_workbook(file_path)

        sheet_name = ConstData.excel_sheet_stats

        try:
            ExcelUtility.write_data_to_sheet(workbook, sheet_name, data, ConstData.excel_cols_repayments, 1)

            workbook.save(file_path)
        except (Exception,):
            raise
        finally:
            MiscUtility.close_workbook(workbook)


    @staticmethod
    def compute_stats(loans: list[Loan], refunded_slices: dict[str, list[(str, str)]]):
        """

        :param loans:
        :param refunded_slices:
        :return:
        """
        previous_loanee: Loanee | None = None
        cells_values: dict[str, ExcelCellData] = {}
        loan_sum = 0

        # Headers appending
        RepaymentUtility.fill_headers_cells(cells_values, ConstData.excel_cols_stats)

        excel_row_index = 0

        for for_index in range((length := len(loans)) + 1):
            loan = loans[for_index] if for_index < length else None

            if previous_loanee is not None and (for_index == length or previous_loanee != loan.loanee):
                excel_row_index += + 1

                # Fill loanee details
                RepaymentUtility.fill_loanee_cells(cells_values, previous_loanee, excel_row_index, ConstData.excel_cols_stats)

                # Total fields
                cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_stats_loan_total,
                                                                               excel_row_index, ConstData.excel_cols_stats)
                cells_values[cell_index] = ExcelCellData(loan_sum)

                # Total refunded fields
                cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_stats_loan_total_refunded,
                                                                    excel_row_index, ConstData.excel_cols_stats)
                refunded_total: int
                loanee_refunded_slices = refunded_slices.get(previous_loanee.ID)

                if loanee_refunded_slices is not None:
                    refunded_total = sum([sl_tpl[2] for sl_tpl in loanee_refunded_slices])
                else:
                    refunded_total = 0

                cells_values[cell_index] = ExcelCellData(refunded_total)

                # Total remaining
                cell_index = ExcelUtility.get_cell_from_column_name(ConstData.excel_col_stats_loan_total_remaining,
                                                                    excel_row_index, ConstData.excel_cols_stats)
                cells_values[cell_index] = ExcelCellData(loan_sum - refunded_total)

                # --
                loan_sum = 0

            if loan is not None:
                loan_sum += loan.amount
                previous_loanee = loan.loanee

        return cells_values
