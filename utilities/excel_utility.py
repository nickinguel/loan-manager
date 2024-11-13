from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from configs.constant_data import ConstData
from models.excel_cell_data import ExcelCellData


class ExcelUtility:

    @staticmethod
    def write_data_to_sheet(workbook: Workbook, sheet_name: str, cells_data: dict[str, ExcelCellData], columns: tuple[str, ...],
                            sheet_index: int = None) -> Worksheet:
        from utilities.repayment_utility import RepaymentUtility

        sheet: Worksheet

        if sheet_name in workbook.sheetnames:
            sheet = workbook.get_sheet_by_name(sheet_name)
        else:
            sheet = workbook.create_sheet(sheet_name, index=sheet_index)

        ExcelUtility.clean_sheet(sheet)

        for cell_index, cell_data in cells_data.items():
            cell: Cell = sheet[cell_index]

            cell.alignment = RepaymentUtility.centered_alignment
            cell.value = cell_data.value

            if cell_data.font:
                cell.font = cell_data.font

        RepaymentUtility.stylize_sheet_general_cells(sheet, columns)

        return sheet

    @staticmethod
    def clean_sheet(sheet: Worksheet):
        for cells_tpl in sheet.iter_rows(min_row=2, values_only=False):
            for cell in cells_tpl:
                cell.value = ""

    @staticmethod
    def get_cell_from_column_name(column_name: str, data_row_number, columns: tuple[str, ...] = ConstData.excel_cols_repayments) -> str | None:
        """
        Given a column name and corresponding data row (ignoring headers row), retrieve the corresponding cell index suc as A1, C4

        :param column_name:
        :param data_row_number:
        :param columns
        :return:
        """

        if column_name not in columns:
            return

        index = "{0}{1}".format(
            ConstData.alphabet[columns.index(column_name)],
            data_row_number + 1
        )

        return index
