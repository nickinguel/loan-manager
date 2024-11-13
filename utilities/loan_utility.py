from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet
from configs.constant_data import ConstData
from models.loan import Loan
from models.loanee import Loanee
from utilities.misc_utility import MiscUtility
import re


class LoanUtility:

    @staticmethod
    def read_loans(file_path: str) -> list[Loan]:
        workbook: Workbook = None
        loans: list[Loan] = []

        try:
            workbook = MiscUtility.open_workbook(file_path)

            if ConstData.excel_sheet_loan not in workbook.sheetnames:
                raise Exception(ConstData.message_loan_sheet_missing.format(file_path))

            loan_sheet = workbook.get_sheet_by_name(ConstData.excel_sheet_loan)

            # Check missing headers
            header_tuple = LoanUtility.check_required_columns_headers(loan_sheet)
            headers = header_tuple[0]
            missing_headers = header_tuple[1]

            if len(missing_headers) > 0:
                raise Exception("Dans la feuille '{0}', les colonnes suivantes sont manquantes : {1}".format(
                    ConstData.excel_sheet_loan,
                    MiscUtility.format_array_as_bullets(missing_headers)
                ))

            # Parse loans
            loans = LoanUtility.parse_loans(loan_sheet, headers)

        except Exception as e:
            raise e
        finally:
            if workbook is not None:
                MiscUtility.close_workbook(workbook)

        return loans

    @staticmethod
    def parse_loans(sheet: Worksheet, headers: dict[str, int]) -> list[Loan]:
        """
        Parse the Loan sheet and convert rows to corresponding :Loan object
        :param sheet:
        :param headers:
        :return:
        """

        loans: list[Loan] = []

        for cells_tpl in sheet.iter_rows(min_row=2, values_only=False):
            cells_values = tuple([cell.value for cell in cells_tpl])
            missing_values = LoanUtility.check_loan_row(headers, cells_tpl)

            if len(missing_values) > 0:
                raise Exception("Dans la feuille '{0}', la ligne '{1}' a des valeurs manquantes ou incorrectes pour les colonnes suivantes : {2}"
                                .format(ConstData.excel_sheet_loan, cells_tpl[0].row.real, MiscUtility.format_array_as_bullets(missing_values))
                            )

            loan = LoanUtility.instantiate_loan_from_xl_row(cells_values, headers)
            loans.append(loan)

        return loans

    loanees: dict[str, Loanee] = None

    @staticmethod
    def instantiate_loan_from_xl_row(cells_values: tuple[str, ...], headers: dict[str, int]) -> Loan:
        """
        Create a Loan object from excel data row
        :param cells_values:
        :param headers:
        :return:
        """

        if LoanUtility.loanees is None:
            LoanUtility.loanees = {}

        loanee_id = cells_values[headers[ConstData.excel_col_loan_debtor_ID]]
        loanee: Loanee = LoanUtility.loanees.get(loanee_id)

        if not loanee:
            loanee = Loanee(
                loanee_id,
                cells_values[headers[ConstData.excel_col_loan_debtor_first_name]],
                cells_values[headers[ConstData.excel_col_loan_debtor_last_name]]
            )

            LoanUtility.loanees[loanee_id] = loanee

        amount = int(cells_values[headers[ConstData.excel_col_loan_amount]])
        repayment_logic_converted = LoanUtility.convert_repayment_logic(cells_values[headers[ConstData.excel_col_loan_repayment_logic]], amount)

        loan = Loan(
            loanee,
            amount,
            repayment_logic_converted,
            from_excel(cells_values[headers[ConstData.excel_col_loan_date]])
        )

        return loan

    @staticmethod
    def check_loan_row(headers: dict[str, int], row_values: tuple[Cell, ...]) -> list[str]:
        """
        Check if some data rows have wrong values (empty or incorrect type)
        :param headers: Headers column indexes
        :param row_values: The values in the row for each column
        :return:
        """

        missing_cells = []
        headers_copy = dict((key, val) for key, val in headers.items() if key not in [
            ConstData.excel_col_loan_amount,
            ConstData.excel_col_loan_date,
            ConstData.excel_col_loan_repayment_logic
        ])

        if re.search("^\\d+$", str(row_values[headers[ConstData.excel_col_loan_amount]].value)) is None:
            missing_cells.append(ConstData.excel_col_loan_amount)

        if re.search(r"^\d+(\s*[+*]\s*\d+)*$", str(row_values[headers[ConstData.excel_col_loan_repayment_logic]].value)) is None:
            missing_cells.append(ConstData.excel_col_loan_repayment_logic)
        else:
            cell: Cell = row_values[headers[ConstData.excel_col_loan_amount]]
            try:
                converted_logic = LoanUtility.convert_repayment_logic(
                    row_values[headers[ConstData.excel_col_loan_repayment_logic]].value,
                    int(cell.value)
                )
            except Exception as e:
                raise Exception(str(e).format(cell.row.real))

            repayment_amount = eval(converted_logic)

            if int(row_values[headers[ConstData.excel_col_loan_amount]].value) != repayment_amount:
                missing_cells.append(ConstData.excel_col_loan_repayment_logic)

        try:
            _ = from_excel(row_values[headers[ConstData.excel_col_loan_date]].value)
        except (Exception,):
            missing_cells.append(ConstData.excel_col_loan_date)

        for col_name, col_index in headers_copy.items():
            if (cell_val := row_values[col_index].value) is None or len(str(cell_val)) == 0:
                missing_cells.append(col_name)

        return missing_cells

    @staticmethod
    def convert_repayment_logic(logic: str, loan_amount: int):
        """
        Takes the original repayment logic from the Excel file and convert it to basic + operation
        Exemple : 1000 * 2 ==> 1000 + 1000
        :param logic:
        :param loan_amount:
        :return:
        """

        logic_str = str(logic)

        if re.search(r"^\d+$", logic_str) is not None:
            slice_amount = int(logic_str)

            if loan_amount % slice_amount != 0:
                raise Exception(f"Le prêt de '{loan_amount}' n'est pas parfaitement divisible par le de remboursement '{slice_amount}' spécifié"
                                + " sur la ligne {0}")

            return LoanUtility.write_slice_n_times(loan_amount // slice_amount, slice_amount)

        while (re_match := re.search(r"(\d+)\s*\*\s*(\d+)", logic_str)) is not None:
            times: int
            amount: int
            multiplication = re_match.group()
            groups = re_match.groups()

            if (first := int(groups[0])) < (second := int(groups[1])):
                times = first
                amount = second
            else:
                times = second
                amount = first

            logic_str = logic_str.replace(multiplication, LoanUtility.write_slice_n_times(times, amount), 1)

        return logic_str

    @staticmethod
    def write_slice_n_times(times: int, amount: any):
        return " + ".join([str(amount) for i in range(times)])

    @staticmethod
    def check_required_columns_headers(sheet: Worksheet):
        """
        Check if some required column headers are missing in the Excel file
        :param sheet:
        :return:
        """
        loan_col_headers = [cell.value for cell in sheet[1]]

        missing_headers = [h for h in ConstData.excel_cols_loan if h not in loan_col_headers]
        headers = dict([(cell.value, index) for index, cell in enumerate(sheet[1]) if cell.value in ConstData.excel_cols_loan])

        return headers, missing_headers

